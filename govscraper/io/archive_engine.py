"""
Core incremental archive logic — library module (no CLI, no checkpoint I/O).

Callers are responsible for checkpoint persistence (file or database).

Exported functions
------------------
run_bootstrap(url, archive_dir, session, name_override, progress_cb)
    -> (file_info: dict, checkpoint: dict)

run_incremental(url, archive_dir, checkpoint, session)
    -> (new_count: int, updated_checkpoint: dict)
"""

import csv
import logging
import os
from datetime import datetime, timezone
from typing import Callable, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from file_handler import sanitize_filename
from scraper_engine import GovILScraper, GovILSession

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# CSV / Excel helpers
# ---------------------------------------------------------------------------

def append_to_csv(csv_path: str, new_items: list, headers: list) -> None:
    """Append rows to an existing CSV. Extra columns in new items are dropped."""
    unknown = set()
    for item in new_items:
        unknown.update(k for k in item if k not in headers)
    if unknown:
        logger.warning(
            "New items have columns not in archive schema (dropped): %s. "
            "Re-bootstrap to capture them.",
            sorted(unknown),
        )
    with open(csv_path, "a", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
        for item in new_items:
            writer.writerow(item)


def regenerate_excel_from_csv(
    csv_path: str, excel_path: str, headers: list, sheet_name: str
) -> None:
    """Rewrite Excel from the master CSV. RTL Hebrew layout."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    ws.sheet_view.rightToLeft = True

    bold = Font(bold=True, size=11)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold
        cell.alignment = Alignment(horizontal="right", wrap_text=True)

    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row_idx, row in enumerate(reader, 2):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(header, "") or "")

    # Column widths: sample first 100 rows for speed
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row_idx in range(2, min(ws.max_row + 1, 102)):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    tmp = excel_path + ".tmp"
    wb.save(tmp)
    os.replace(tmp, excel_path)
    logger.info("Excel regenerated: %s (%d rows)", excel_path, ws.max_row - 1)


# ---------------------------------------------------------------------------
# Bootstrap (first run — full scrape)
# ---------------------------------------------------------------------------

def run_bootstrap(
    url: str,
    archive_dir: str,
    session: GovILSession,
    name_override: str = "",
    progress_cb: Optional[Callable] = None,
) -> Tuple[dict, dict]:
    """Full scrape, write master CSV + Excel.

    Returns
    -------
    file_info : dict
        csv_path, excel_path, csv_basename, excel_basename,
        collector_name, record_count, total_count, column_headers, warning
    checkpoint : dict
        State dict for subsequent incremental runs. ``known_urls`` is a set.
    """
    logger.info("BOOTSTRAP: full scrape of %s", url)

    def _default_cb(**kw):
        msg = kw.get("message", "")
        total = kw.get("total", 0)
        current = kw.get("current", 0)
        if total > 0:
            logger.info("  [%d/%d] %s", current, total, msg)

    scraper = GovILScraper(session, progress_callback=progress_cb or _default_cb)
    result = scraper.scrape(url)

    if not result.items:
        raise RuntimeError("Bootstrap scrape returned 0 items — aborting")

    logger.info("Bootstrap: %d items, %d columns", len(result.items), len(result.column_headers))

    os.makedirs(archive_dir, exist_ok=True)
    safe_name = sanitize_filename(name_override or result.collector_name or "archive")
    csv_path = os.path.join(archive_dir, f"{safe_name}.csv")
    excel_path = os.path.join(archive_dir, f"{safe_name}.xlsx")

    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=result.column_headers, extrasaction="ignore")
        writer.writeheader()
        for item in result.items:
            writer.writerow(item)
    logger.info("CSV written: %s (%d rows)", csv_path, len(result.items))

    regenerate_excel_from_csv(csv_path, excel_path, result.column_headers, safe_name)

    known_urls = {item.get("url", "") for item in result.items if item.get("url")}

    file_info = {
        "csv_path": csv_path,
        "excel_path": excel_path,
        "csv_basename": os.path.basename(csv_path),
        "excel_basename": os.path.basename(excel_path),
        "collector_name": result.collector_name,
        "record_count": len(result.items),
        "total_count": result.total_count,
        "column_headers": result.column_headers,
        "warning": result.warning or "",
    }

    checkpoint = {
        "url": url,
        "last_known_total": result.total_count,
        "last_run_utc": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "last_run_new_items": len(result.items),
        "known_urls": known_urls,
        "archive_csv": os.path.basename(csv_path),
        "archive_excel": os.path.basename(excel_path),
        "column_headers": result.column_headers,
        "total_archived": len(result.items),
    }

    logger.info("Bootstrap complete. total=%d", result.total_count)
    return file_info, checkpoint


# ---------------------------------------------------------------------------
# Incremental update (daily runs)
# ---------------------------------------------------------------------------

def run_incremental(
    url: str,
    archive_dir: str,
    checkpoint: dict,
    session: GovILSession,
) -> Tuple[int, dict]:
    """Fetch only new items and append them to the master CSV.

    Parameters
    ----------
    checkpoint : dict
        Must contain ``known_urls`` as a set (not a list).

    Returns
    -------
    new_count : int
    updated_checkpoint : dict   (``known_urls`` remains a set)
    """
    scraper = GovILScraper(session)
    last_total = checkpoint["last_known_total"]

    logger.info("Checking current total (last known: %d)...", last_total)
    current_total, _ = scraper.fetch_traditional_page(url, skip=0, limit=1)
    logger.info("Current total: %d", current_total)

    if current_total < last_total:
        logger.info(
            "API total decreased from %d to %d (items removed). Archive preserved.",
            last_total, current_total,
        )
        checkpoint = dict(checkpoint)
        checkpoint["last_known_total"] = current_total
        checkpoint["last_run_utc"] = datetime.now(timezone.utc).isoformat(timespec="seconds")
        checkpoint["last_run_new_items"] = 0
        return 0, checkpoint

    if current_total == last_total:
        logger.info("No new items — archive is up to date.")
        checkpoint = dict(checkpoint)
        checkpoint["last_run_utc"] = datetime.now(timezone.utc).isoformat(timespec="seconds")
        checkpoint["last_run_new_items"] = 0
        return 0, checkpoint

    delta = current_total - last_total
    logger.info("Delta: %d new items.", delta)

    fetch_limit = delta + max(5, int(delta * 0.2))
    logger.info("Fetching %d items (delta + buffer)...", fetch_limit)
    _, raw_items = scraper.fetch_traditional_page(url, skip=0, limit=fetch_limit)

    if not raw_items:
        logger.warning("API returned 0 items despite delta=%d. Skipping.", delta)
        return 0, checkpoint

    flat_items = [scraper._flatten_item(item) for item in raw_items]

    known_urls = set(checkpoint.get("known_urls") or set())
    new_items = []
    for item in flat_items:
        item_url = item.get("url", "")
        if item_url and item_url in known_urls:
            continue
        new_items.append(item)
        if item_url:
            known_urls.add(item_url)

    logger.info("After deduplication: %d genuinely new items.", len(new_items))

    checkpoint = dict(checkpoint)
    checkpoint["last_known_total"] = current_total
    checkpoint["last_run_utc"] = datetime.now(timezone.utc).isoformat(timespec="seconds")
    checkpoint["known_urls"] = known_urls

    if not new_items:
        logger.info("All fetched items already archived (reorder). Updating baseline.")
        checkpoint["last_run_new_items"] = 0
        return 0, checkpoint

    headers = checkpoint["column_headers"]
    csv_path = os.path.join(archive_dir, checkpoint["archive_csv"])
    excel_path = os.path.join(archive_dir, checkpoint["archive_excel"])
    sheet_name = os.path.splitext(checkpoint["archive_csv"])[0]

    append_to_csv(csv_path, new_items, headers)
    regenerate_excel_from_csv(csv_path, excel_path, headers, sheet_name)

    checkpoint["last_run_new_items"] = len(new_items)
    checkpoint["total_archived"] = checkpoint.get("total_archived", 0) + len(new_items)

    logger.info(
        "Done. Added %d new items. Archive total: %d.",
        len(new_items), checkpoint["total_archived"],
    )
    return len(new_items), checkpoint
