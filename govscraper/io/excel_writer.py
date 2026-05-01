"""Excel writer — RTL Hebrew layout, auto-fit columns."""
from __future__ import annotations

import logging
import os
from typing import Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from .sanitize import sanitize_filename

logger = logging.getLogger(__name__)


def write(
    output_dir: str,
    name: str,
    rows: Iterable[dict],
    columns: Sequence[str],
) -> str:
    """Write `rows` to `<output_dir>/<sanitized name>.xlsx`. Returns absolute path.

    Excel sheet is set to right-to-left (Hebrew). Header row is bolded,
    aligned-right, and wrap-text enabled. Column widths auto-fit based on
    the longer of header text or first 100 row values, capped at 50 chars.
    """
    filename = sanitize_filename(name) + ".xlsx"
    filepath = os.path.join(output_dir, filename)

    rows_list = list(rows)
    columns = list(columns)

    wb = Workbook()
    ws = wb.active
    ws.title = name[:31]  # Excel sheet-name limit
    ws.sheet_view.rightToLeft = True

    header_font = Font(bold=True, size=11)
    for col_idx, header in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="right", wrap_text=True)

    for row_idx, item in enumerate(rows_list, 2):
        for col_idx, header in enumerate(columns, 1):
            value = item.get(header, "")
            if value is None:
                value = ""
            ws.cell(row=row_idx, column=col_idx, value=str(value))

    # Auto-fit column widths (sample first 100 rows for speed)
    sample_limit = min(len(rows_list) + 2, 102)
    for col_idx, header in enumerate(columns, 1):
        max_len = len(str(header))
        for row_idx in range(2, sample_limit):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    wb.save(filepath)
    logger.info("Excel exported: %s (%d rows)", filepath, len(rows_list))
    return filepath
