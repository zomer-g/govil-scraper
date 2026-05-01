"""
File Handler — CSV/Excel/GeoJSON export, attachment download, ZIP packaging.
All exports use utf-8-sig for Hebrew compatibility in Excel.
"""

import csv
import json
import os
import re
import logging
import tempfile
import zipfile
from datetime import datetime
from typing import List, Optional, Callable

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from govscraper.scrapers.govil.legacy_engine import GovILSession, ScrapeResult, FileAttachment
from govscraper.io.sanitize import sanitize_filename

logger = logging.getLogger(__name__)


class FileHandler:
    """Handles CSV/Excel export, file downloads, and ZIP packaging."""

    def __init__(self, session: GovILSession, output_dir: Optional[str] = None):
        self.session = session
        self.output_dir = output_dir or tempfile.mkdtemp(prefix="govil_")
        os.makedirs(self.output_dir, exist_ok=True)

    def export_csv(self, result: ScrapeResult) -> str:
        """Export scraped data to CSV with utf-8-sig encoding."""
        filename = sanitize_filename(result.collector_name) + ".csv"
        filepath = os.path.join(self.output_dir, filename)

        with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(
                f,
                fieldnames=result.column_headers,
                extrasaction="ignore",
            )
            writer.writeheader()
            for item in result.items:
                writer.writerow(item)

        logger.info("CSV exported: %s (%d rows)", filepath, len(result.items))
        return filepath

    def export_excel(self, result: ScrapeResult) -> str:
        """Export scraped data to Excel with RTL support."""
        filename = sanitize_filename(result.collector_name) + ".xlsx"
        filepath = os.path.join(self.output_dir, filename)

        wb = Workbook()
        ws = wb.active
        ws.title = result.collector_name[:31]  # Excel sheet name limit
        ws.sheet_view.rightToLeft = True

        # Header row
        header_font = Font(bold=True, size=11)
        for col_idx, header in enumerate(result.column_headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="right", wrap_text=True)

        # Data rows
        for row_idx, item in enumerate(result.items, 2):
            for col_idx, header in enumerate(result.column_headers, 1):
                value = item.get(header, "")
                if value is None:
                    value = ""
                ws.cell(row=row_idx, column=col_idx, value=str(value))

        # Auto-fit column widths
        for col_idx, header in enumerate(result.column_headers, 1):
            max_len = len(str(header))
            for row_idx in range(2, min(len(result.items) + 2, 102)):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val:
                    max_len = max(max_len, len(str(cell_val)))
            adjusted = min(max_len + 2, 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted

        wb.save(filepath)
        logger.info("Excel exported: %s (%d rows)", filepath, len(result.items))
        return filepath

    def download_attachments(
        self,
        attachments: List[FileAttachment],
        progress_callback: Optional[Callable] = None,
        skip_existing: bool = False,
    ) -> List[str]:
        """Download all file attachments. Returns list of local file paths.

        If skip_existing=True, files that already exist on disk (with size > 0)
        are skipped — useful for local scraping to avoid re-downloading.
        """
        if not attachments:
            return []

        att_dir = os.path.join(self.output_dir, "attachments")
        os.makedirs(att_dir, exist_ok=True)

        paths = []
        used_names = set()
        skipped = 0

        for idx, att in enumerate(attachments):
            try:
                # Determine the canonical filename first
                filename = sanitize_filename(att.filename)
                if not filename or filename == "unnamed":
                    filename = att.url.split("/")[-1].split("?")[0] or "file"
                    filename = sanitize_filename(filename)

                dest_path = os.path.join(att_dir, filename)

                # Skip if file already exists and has content
                if skip_existing and os.path.exists(dest_path) and os.path.getsize(dest_path) > 0:
                    paths.append(dest_path)
                    used_names.add(filename.lower())
                    skipped += 1
                else:
                    local_path = self._download_single(att, att_dir, used_names)
                    paths.append(local_path)
            except Exception as e:
                logger.warning("Failed to download %s: %s", att.url, e)

            if progress_callback:
                progress_callback(
                    current=idx + 1,
                    total=len(attachments),
                    message=f"הורדו {idx + 1} מתוך {len(attachments)} קבצים",
                )

        if skipped:
            logger.info("Skipped %d existing files", skipped)
        logger.info("Downloaded %d / %d attachments (%d skipped)",
                     len(paths) - skipped, len(attachments), skipped)
        return paths

    def _download_single(
        self, att: FileAttachment, dest_dir: str, used_names: set
    ) -> str:
        """Download a single file attachment.

        Uses download_file() which strips API-specific headers (Origin,
        Referer, Accept) for external domains — those headers cause servers
        like police.gov.il to reject requests.
        """
        resp = self.session.download_file(att.url)

        if not resp.content:
            raise RuntimeError(f"Empty response for {att.url}")

        filename = sanitize_filename(att.filename)
        if not filename or filename == "unnamed":
            filename = att.url.split("/")[-1].split("?")[0] or "file"
            filename = sanitize_filename(filename)

        # Deduplicate names
        base, ext = os.path.splitext(filename)
        candidate = filename
        counter = 1
        while candidate.lower() in used_names:
            candidate = f"{base}_{counter}{ext}"
            counter += 1
        used_names.add(candidate.lower())

        filepath = os.path.join(dest_dir, candidate)
        with open(filepath, "wb") as f:
            f.write(resp.content)

        logger.debug("Downloaded: %s (%d bytes)", filepath, len(resp.content))
        return filepath

    def create_zip(
        self,
        csv_path: str,
        excel_path: str,
        attachment_paths: List[str],
    ) -> str:
        """Package everything into a ZIP file."""
        zip_name = sanitize_filename(
            os.path.basename(csv_path).replace(".csv", "")
        ) + ".zip"
        zip_path = os.path.join(self.output_dir, zip_name)

        folder_name = os.path.basename(csv_path).replace(".csv", "")

        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            # Data files at root of folder
            zf.write(csv_path, f"{folder_name}/{os.path.basename(csv_path)}")
            zf.write(excel_path, f"{folder_name}/{os.path.basename(excel_path)}")

            # Attachments in subfolder
            for path in attachment_paths:
                arc_name = f"{folder_name}/attachments/{os.path.basename(path)}"
                zf.write(path, arc_name)

        logger.info("ZIP created: %s", zip_path)
        return zip_path

    def get_all_file_records(self, base_dir: str) -> List[dict]:
        """Scan output_dir and return metadata for all files.

        Returns list of dicts with: filename, file_type, category,
        size_bytes, rel_path (relative to base_dir).
        Used to populate the files table in the database.
        """
        records = []
        for root, _dirs, files in os.walk(self.output_dir):
            for fname in files:
                full = os.path.join(root, fname)
                rel = os.path.relpath(full, base_dir).replace("\\", "/")
                ext = os.path.splitext(fname)[1].lower().lstrip(".")

                if ext == "csv":
                    category = "csv"
                elif ext == "xlsx":
                    category = "excel"
                elif ext == "geojson":
                    category = "geojson"
                elif ext == "json" and fname == "manifest.json":
                    category = "manifest"
                else:
                    category = "attachment"

                records.append({
                    "filename": fname,
                    "file_type": ext,
                    "category": category,
                    "size_bytes": os.path.getsize(full),
                    "rel_path": rel,
                })
        return records

    # ---- GovMap-only exports (no-op for non-GovMap results) -------------

    def export_geojson(self, result: ScrapeResult) -> str:
        """Emit a FeatureCollection in WGS84.

        Reads result.features (populated by govmap_engine for GOVMAP_LAYER
        scrapes; empty list for other page types so this method is safe to
        call unconditionally — but a no-op file is still written, which is
        rarely what you want; check `result.features` before calling).
        """
        filename = sanitize_filename(result.collector_name) + ".geojson"
        filepath = os.path.join(self.output_dir, filename)

        fc = {
            "type": "FeatureCollection",
            "name": result.collector_name,
            "crs": {
                "type": "name",
                "properties": {"name": "urn:ogc:def:crs:OGC:1.3:CRS84"},
            },
            "metadata": {
                "layer_id": getattr(result, "layer_id", ""),
                "bbox_itm": list(result.bbox_itm) if result.bbox_itm else None,
                "bbox_wgs84": list(result.bbox_wgs84) if result.bbox_wgs84 else None,
                "geometry_type": getattr(result, "geometry_type", ""),
                "feature_count": result.total_count,
                "scraped_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
            },
            "features": [
                {"type": "Feature",
                 "geometry": f.geometry,
                 "properties": f.properties}
                for f in (result.features or [])
            ],
        }
        with open(filepath, "w", encoding="utf-8") as fh:
            json.dump(fc, fh, ensure_ascii=False)
        logger.info("GeoJSON exported: %s (%d features)",
                    filepath, len(result.features or []))
        return filepath

    def write_manifest(self, result: ScrapeResult, files: dict) -> str:
        """Write a sidecar manifest.json the upload endpoint reads to
        populate geo columns in the collections table.
        """
        path = os.path.join(self.output_dir, "manifest.json")
        manifest = {
            "layer_id": getattr(result, "layer_id", ""),
            "collector_name": result.collector_name,
            "page_type": (result.page_type.value
                          if hasattr(result.page_type, "value")
                          else str(result.page_type)),
            "bbox_itm": list(result.bbox_itm) if result.bbox_itm else None,
            "bbox_wgs84": list(result.bbox_wgs84) if result.bbox_wgs84 else None,
            "geometry_type": getattr(result, "geometry_type", ""),
            "feature_count": result.total_count,
            "srs": getattr(result, "srs", ""),
            "warning": result.warning or "",
            "files": {k: os.path.relpath(v, self.output_dir).replace("\\", "/")
                      for k, v in files.items() if v},
        }
        with open(path, "w", encoding="utf-8") as fh:
            json.dump(manifest, fh, ensure_ascii=False, indent=2)
        return path
