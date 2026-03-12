"""
File Handler — CSV/Excel export, attachment download, ZIP packaging.
All exports use utf-8-sig for Hebrew compatibility in Excel.
"""

import csv
import os
import re
import logging
import tempfile
import zipfile
from typing import List, Optional, Callable

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from scraper_engine import GovILSession, ScrapeResult, FileAttachment

logger = logging.getLogger(__name__)


def sanitize_filename(name: str, max_len: int = 200) -> str:
    """Remove invalid filesystem characters and truncate."""
    if not name:
        return "unnamed"
    # Remove chars invalid on Windows
    name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', '_', name)
    # Collapse multiple underscores
    name = re.sub(r'_+', '_', name).strip('_. ')
    if len(name) > max_len:
        base, ext = os.path.splitext(name)
        name = base[:max_len - len(ext)] + ext
    return name or "unnamed"


class FileHandler:
    """Handles CSV/Excel export, file downloads, and ZIP packaging."""

    def __init__(self, session: GovILSession, output_dir: Optional[str] = None):
        self.session = session
        self.output_dir = output_dir or tempfile.mkdtemp(prefix="govil_")
        os.makedirs(self.output_dir, exist_ok=True)

    def export_csv(self, result: ScrapeResult) -> str:
        """Export scraped data to CSV with utf-8-sig encoding."""
        filename = sanitize_filename(result.collector_name) + ".csv"
        filepath = os.path.join(self.output_dir, filename)

        with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(
                f,
                fieldnames=result.column_headers,
                extrasaction="ignore",
            )
            writer.writeheader()
            for item in result.items:
                writer.writerow(item)

        logger.info("CSV exported: %s (%d rows)", filepath, len(result.items))
        return filepath

    def export_excel(self, result: ScrapeResult) -> str:
        """Export scraped data to Excel with RTL support."""
        filename = sanitize_filename(result.collector_name) + ".xlsx"
        filepath = os.path.join(self.output_dir, filename)

        wb = Workbook()
        ws = wb.active
        ws.title = result.collector_name[:31]  # Excel sheet name limit
        ws.sheet_view.rightToLeft = True

        # Header row
        header_font = Font(bold=True, size=11)
        for col_idx, header in enumerate(result.column_headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="right", wrap_text=True)

        # Data rows
        for row_idx, item in enumerate(result.items, 2):
            for col_idx, header in enumerate(result.column_headers, 1):
                value = item.get(header, "")
                if value is None:
                    value = ""
                ws.cell(row=row_idx, column=col_idx, value=str(value))

        # Auto-fit column widths
        for col_idx, header in enumerate(result.column_headers, 1):
            max_len = len(str(header))
            for row_idx in range(2, min(len(result.items) + 2, 102)):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val:
                    max_len = max(max_len, len(str(cell_val)))
            adjusted = min(max_len + 2, 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted

        wb.save(filepath)
        logger.info("Excel exported: %s (%d rows)", filepath, len(result.items))
        return filepath

    def download_attachments(
        self,
        attachments: List[FileAttachment],
        progress_callback: Optional[Callable] = None,
    ) -> List[str]:
        """Download all file attachments. Returns list of local file paths."""
        if not attachments:
            return []

        att_dir = os.path.join(self.output_dir, "attachments")
        os.makedirs(att_dir, exist_ok=True)

        paths = []
        used_names = set()

        for idx, att in enumerate(attachments):
            try:
                local_path = self._download_single(att, att_dir, used_names)
                paths.append(local_path)
            except Exception as e:
                logger.warning("Failed to download %s: %s", att.url, e)

            if progress_callback:
                progress_callback(
                    current=idx + 1,
                    total=len(attachments),
                    message=f"הורדו {idx + 1} מתוך {len(attachments)} קבצים",
                )

        logger.info("Downloaded %d / %d attachments", len(paths), len(attachments))
        return paths

    def _download_single(
        self, att: FileAttachment, dest_dir: str, used_names: set
    ) -> str:
        """Download a single file attachment.

        Uses non-streaming mode so cloudscraper can properly handle
        Cloudflare challenges (stream=True skips challenge solving).
        """
        resp = self.session.get(att.url, stream=False)

        if not resp.content:
            raise RuntimeError(f"Empty response for {att.url}")

        filename = sanitize_filename(att.filename)
        if not filename or filename == "unnamed":
            filename = att.url.split("/")[-1].split("?")[0] or "file"
            filename = sanitize_filename(filename)

        # Deduplicate names
        base, ext = os.path.splitext(filename)
        candidate = filename
        counter = 1
        while candidate.lower() in used_names:
            candidate = f"{base}_{counter}{ext}"
            counter += 1
        used_names.add(candidate.lower())

        filepath = os.path.join(dest_dir, candidate)
        with open(filepath, "wb") as f:
            f.write(resp.content)

        logger.debug("Downloaded: %s (%d bytes)", filepath, len(resp.content))
        return filepath

    def create_zip(
        self,
        csv_path: str,
        excel_path: str,
        attachment_paths: List[str],
    ) -> str:
        """Package everything into a ZIP file."""
        zip_name = sanitize_filename(
            os.path.basename(csv_path).replace(".csv", "")
        ) + ".zip"
        zip_path = os.path.join(self.output_dir, zip_name)

        folder_name = os.path.basename(csv_path).replace(".csv", "")

        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            # Data files at root of folder
            zf.write(csv_path, f"{folder_name}/{os.path.basename(csv_path)}")
            zf.write(excel_path, f"{folder_name}/{os.path.basename(excel_path)}")

            # Attachments in subfolder
            for path in attachment_paths:
                arc_name = f"{folder_name}/attachments/{os.path.basename(path)}"
                zf.write(path, arc_name)

        logger.info("ZIP created: %s", zip_path)
        return zip_path
